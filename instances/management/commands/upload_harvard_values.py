from django.core.management.base import BaseCommand, CommandError
from milestones.models import Milestone, MilestoneAreaValue
from openpyxl import Workbook, load_workbook
from content_manager import settings
import os

class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        try:

            archivo = load_workbook(os.path.join(settings.BASE_DIR, 'instances/milestones_area_values.xlsx'))
            sheet = archivo.active

            for row in sheet.iter_rows():
                milestone_code = row[0].value
                if(Milestone.objects.filter(code=milestone_code).exists()):
                    self.stdout.write(f"Milestone code: {milestone_code}")
                    milestone = Milestone.objects.get(code=milestone_code)
                    valores = [
                        {'area_id': 2, 'valor': row[1].value},
                        {'area_id': 1, 'valor': row[2].value},
                        {'area_id': 43, 'valor': row[3].value},
                        {'area_id': 3, 'valor': row[4].value},
                    ]

                    for v in valores:
                        area_value = MilestoneAreaValue.objects.get_or_create(
                            value=v['valor'],
                            area_id=v['area_id'],
                            milestone_id=milestone.id
                        )

            self.stdout.write('finished!')

        except Exception as err:
            self.stdout.write(f'Ha ocurrido un error {err}')
